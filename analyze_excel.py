import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys

# Set UTF-8 encoding for stdout
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

def analyze_excel(filepath, name):
    print(f"\n{'='*80}")
    print(f"Excel: {name}")
    print(f"File: {filepath}")
    print(f"{'='*80}")

    if not os.path.exists(filepath):
        print(f"ERROR: File not found!")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    print(f"Sheet name: {ws.title}")
    print(f"Total rows: {ws.max_row}")
    print(f"Total columns: {ws.max_column}")

    # Print first 20 rows to identify headers
    print(f"\nFirst 20 rows (raw data):")
    print("-" * 80)

    header_row = None
    for i, row in enumerate(list(ws.iter_rows(max_row=20)), start=1):
        row_values = [cell.value for cell in row]
        print(f"Row {i}: {row_values}")

        # Try to detect header row (row with most non-None string values)
        if header_row is None and any(row_values):
            non_none_count = sum(1 for v in row_values if v is not None and isinstance(v, str) and v.strip())
            if non_none_count >= 3:  # At least 3 non-empty string values
                if header_row is None:
                    header_row = i

    # Identify header row more intelligently
    print(f"\n\n{'='*80}")
    print("HEADER ANALYSIS:")
    print(f"{'='*80}")

    if header_row is None:
        header_row = 1

    print(f"Detected Header Row: {header_row}")

    # Get headers
    headers = []
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    for idx, cell in enumerate(header_cells):
        col_letter = get_column_letter(idx + 1)
        header_value = cell.value if cell.value is not None else ""
        headers.append((col_letter, header_value))

    print(f"\nHeaders (in order):")
    for col_letter, header_value in headers:
        if header_value:  # Only print non-empty headers
            print(f"  Column {col_letter}: {header_value}")

    # Get sample data rows (next 5 rows after header)
    print(f"\nSample Data Rows (rows {header_row+1} to {header_row+5}):")
    print("-" * 80)

    for i in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        row_data = list(ws.iter_rows(min_row=i, max_row=i))[0]
        row_values = [cell.value for cell in row_data]
        print(f"Row {i}: {row_values}")

        # Show data types for first data row
        if i == header_row + 1:
            print(f"\nData types for first data row (Row {i}):")
            for idx, (col_letter, header_value) in enumerate(headers):
                if header_value and idx < len(row_values):
                    value = row_values[idx]
                    if value is None:
                        dtype = "empty"
                    elif isinstance(value, str):
                        dtype = "string"
                    elif isinstance(value, int):
                        dtype = "integer"
                    elif isinstance(value, float):
                        dtype = "number"
                    elif hasattr(value, 'strftime'):  # datetime
                        dtype = "date/datetime"
                    else:
                        dtype = type(value).__name__
                    print(f"  Column {col_letter} ({header_value}): {dtype} = {value}")

    wb.close()

# Analyze all three files
files = [
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Invoices\Exaple-Invoices.xlsx", "Invoices"),
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Даночен биланс\РД-Данок на добивка-2024-Example.xlsx", "Даночен биланс (Tax Balance)"),
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Плати\РД-Трошоци за вработени-Example.xlsx", "Плати (Salaries)")
]

for filepath, name in files:
    try:
        analyze_excel(filepath, name)
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"ERROR analyzing {name}:")
        print(f"{'='*80}")
        print(f"{type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
