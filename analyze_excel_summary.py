import openpyxl
from openpyxl.utils import get_column_letter
import sys

# Set UTF-8 encoding for stdout
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

def get_type_name(value):
    """Get a readable type name for a value"""
    if value is None:
        return "empty"
    elif isinstance(value, str):
        return "string"
    elif isinstance(value, int):
        return "integer"
    elif isinstance(value, float):
        return "number"
    elif hasattr(value, 'strftime'):
        return "date/datetime"
    else:
        return type(value).__name__

def analyze_excel_headers(filepath, name):
    print(f"\n{'='*80}")
    print(f"Excel: {name}")
    print(f"File: {filepath}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    print(f"Sheet Name: {ws.title}")
    print(f"Total Rows: {ws.max_row}")
    print(f"Total Columns: {ws.max_column}\n")

    # Show first 12 rows to help identify header row
    print("First 12 rows (to identify header location):")
    print("-" * 80)
    for i in range(1, min(13, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=i, max_row=i))[0]
        # Only show first 10 columns for readability
        values = [str(cell.value)[:50] if cell.value else '' for cell in row[:10]]
        print(f"Row {i:2d}: {values[:3]} {'...' if len(values) > 3 else ''}")

    # Try to detect header row
    print(f"\n{'='*80}")
    print("HEADER DETECTION:")
    print(f"{'='*80}")

    header_row = None
    for i in range(1, min(15, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=i, max_row=i))[0]
        non_empty = sum(1 for cell in row if cell.value and str(cell.value).strip())
        if non_empty >= 3:
            # Check if values look like headers (strings, not too long)
            looks_like_header = True
            header_candidates = []
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip()
                    header_candidates.append(val)
                    # Headers are usually short and don't contain lots of numbers
                    if len(val) > 100 or val.count('\n') > 2:
                        looks_like_header = False

            if looks_like_header and non_empty >= 3:
                print(f"Row {i} looks like headers ({non_empty} non-empty cells)")
                print(f"  Sample headers: {header_candidates[:5]}")
                if header_row is None:
                    header_row = i

    if not header_row:
        header_row = 1
        print("Could not detect header row, using Row 1")

    print(f"\n>>> DETECTED HEADER ROW: {header_row} <<<\n")

    # Get all headers
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    headers = []
    for idx, cell in enumerate(header_cells):
        if cell.value:
            headers.append({
                'index': idx,
                'column': get_column_letter(idx + 1),
                'name': str(cell.value)
            })

    print(f"Headers (in order) - Total: {len(headers)} columns:")
    print("-" * 80)
    for h in headers:
        print(f"  Column {h['column']:3s} (index {h['index']:2d}): {h['name']}")

    # Get first data row
    print(f"\nSample Data Row (Row {header_row + 1}):")
    print("-" * 80)
    data_row = list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1))[0]

    for h in headers:
        value = data_row[h['index']].value
        dtype = get_type_name(value)
        # Truncate long values
        display_value = str(value)[:100] if value else 'None'
        print(f"  Column {h['column']} - {h['name']}:")
        print(f"    Type: {dtype}")
        print(f"    Value: {display_value}")

    # Get second data row too
    if ws.max_row > header_row + 1:
        print(f"\nSample Data Row 2 (Row {header_row + 2}):")
        print("-" * 80)
        data_row2 = list(ws.iter_rows(min_row=header_row + 2, max_row=header_row + 2))[0]

        for h in headers[:5]:  # Just show first 5 for brevity
            value = data_row2[h['index']].value
            dtype = get_type_name(value)
            display_value = str(value)[:100] if value else 'None'
            print(f"  Column {h['column']} - {h['name']}: {dtype} = {display_value}")

    wb.close()

# Analyze the three files
files = [
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Invoices\Exaple-Invoices.xlsx", "Invoices"),
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Даночен биланс\РД-Данок на добивка-2024-Example.xlsx", "Даночен биланс (Tax Balance)"),
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Плати\РД-Трошоци за вработени-Example.xlsx", "Плати (Salaries)")
]

for filepath, name in files:
    try:
        analyze_excel_headers(filepath, name)
    except Exception as e:
        print(f"\nERROR analyzing {name}: {e}")
        import traceback
        traceback.print_exc()

print(f"\n\n{'='*80}")
print("ANALYSIS COMPLETE")
print(f"{'='*80}")
