import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
import json

# Set UTF-8 encoding for stdout
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

def analyze_excel(filepath, name):
    result = {
        'name': name,
        'filepath': filepath,
        'exists': False,
        'error': None,
        'sheet_name': None,
        'total_rows': 0,
        'total_columns': 0,
        'header_row': None,
        'headers': [],
        'first_20_rows': [],
        'sample_data': []
    }

    if not os.path.exists(filepath):
        result['error'] = 'File not found'
        return result

    try:
        result['exists'] = True
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        result['sheet_name'] = ws.title
        result['total_rows'] = ws.max_row
        result['total_columns'] = ws.max_column

        # Get first 20 rows
        for i in range(1, min(21, ws.max_row + 1)):
            row_data = list(ws.iter_rows(min_row=i, max_row=i))[0]
            row_values = []
            for cell in row_data:
                value = cell.value
                if value is None:
                    row_values.append(None)
                elif hasattr(value, 'strftime'):  # datetime
                    row_values.append(value.strftime('%Y-%m-%d'))
                else:
                    row_values.append(str(value))
            result['first_20_rows'].append({
                'row_num': i,
                'values': row_values
            })

        # Detect header row - look for row with most non-None string values
        header_row = None
        max_headers = 0
        for i in range(1, min(15, ws.max_row + 1)):
            row = result['first_20_rows'][i-1]['values'] if i <= len(result['first_20_rows']) else []
            non_none_count = sum(1 for v in row if v is not None and (v.strip() if isinstance(v, str) else True))
            if non_none_count > max_headers and non_none_count >= 3:
                max_headers = non_none_count
                header_row = i

        if header_row is None:
            header_row = 1

        result['header_row'] = header_row

        # Get headers
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
        for idx, cell in enumerate(header_cells):
            col_letter = get_column_letter(idx + 1)
            header_value = cell.value if cell.value is not None else ""
            if header_value:  # Only include non-empty headers
                result['headers'].append({
                    'column': col_letter,
                    'name': str(header_value),
                    'index': idx
                })

        # Get sample data (next 3 rows after header)
        for i in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
            row_data = list(ws.iter_rows(min_row=i, max_row=i))[0]
            row_dict = {}
            for idx, cell in enumerate(row_data):
                value = cell.value
                col_letter = get_column_letter(idx + 1)

                # Determine data type
                if value is None:
                    dtype = "empty"
                    display_value = None
                elif isinstance(value, str):
                    dtype = "string"
                    display_value = value
                elif isinstance(value, int):
                    dtype = "integer"
                    display_value = value
                elif isinstance(value, float):
                    dtype = "number"
                    display_value = value
                elif hasattr(value, 'strftime'):  # datetime
                    dtype = "date/datetime"
                    display_value = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    dtype = type(value).__name__
                    display_value = str(value)

                # Find header name for this column
                header_name = None
                for h in result['headers']:
                    if h['index'] == idx:
                        header_name = h['name']
                        break

                if header_name:
                    row_dict[col_letter] = {
                        'header': header_name,
                        'value': display_value,
                        'type': dtype
                    }

            result['sample_data'].append({
                'row_num': i,
                'data': row_dict
            })

        wb.close()
    except Exception as e:
        result['error'] = f"{type(e).__name__}: {str(e)}"

    return result

# Analyze all three files
files = [
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Invoices\Exaple-Invoices.xlsx", "Invoices"),
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Даночен биланс\РД-Данок на добивка-2024-Example.xlsx", "Даночен биланс (Tax Balance)"),
    (r"C:\BuildApps\document-scanner-desktop\example\Примери за автоматизирање на процеси\Плати\РД-Трошоци за вработени-Example.xlsx", "Плати (Salaries)")
]

all_results = []
for filepath, name in files:
    result = analyze_excel(filepath, name)
    all_results.append(result)

# Save to JSON file
output_file = r"C:\BuildApps\document-scanner-desktop\excel_analysis_results.json"
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(all_results, f, ensure_ascii=False, indent=2)

print(f"Analysis complete! Results saved to: {output_file}")
print("\n" + "="*80)

# Print summary to console
for result in all_results:
    print(f"\n{'='*80}")
    print(f"Excel: {result['name']}")
    print(f"{'='*80}")
    if result['error']:
        print(f"ERROR: {result['error']}")
        continue

    print(f"Sheet: {result['sheet_name']}")
    print(f"Total rows: {result['total_rows']}")
    print(f"Total columns: {result['total_columns']}")
    print(f"Header Row: {result['header_row']}")
    print(f"\nHeaders ({len(result['headers'])} columns):")
    for h in result['headers']:
        print(f"  Column {h['column']}: {h['name']}")

    print(f"\nSample Data (first data row after header):")
    if result['sample_data']:
        first_row = result['sample_data'][0]
        print(f"  Row {first_row['row_num']}:")
        for col, info in first_row['data'].items():
            print(f"    {col} ({info['header']}): {info['type']} = {info['value']}")

print("\n" + "="*80)
print(f"Full results saved to: {output_file}")
